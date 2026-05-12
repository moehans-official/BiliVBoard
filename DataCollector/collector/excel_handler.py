from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
RANK_FILL = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
RANK_FONT = Font(bold=True, color='FFFFFF')


def apply_header_style(ws, headers):
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def apply_center_alignment(ws, row, max_col):
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')


def create_rank_section(ws, sorted_results, data_row_count):
    ws.cell(row=data_row_count + 3, column=1, value='排行榜').font = Font(bold=True, size=12)
    rank_headers = ['排名', 'BVID', '标题', 'UP主', '综合评分']
    rank_start = data_row_count + 4
    
    for col, header in enumerate(rank_headers, start=1):
        cell = ws.cell(row=rank_start, column=col, value=header)
        cell.font = RANK_FONT
        cell.fill = RANK_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for i, data in enumerate(sorted_results, start=1):
        row_num = rank_start + i
        ws.cell(row=row_num, column=1, value=i).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=2, value=data['bvid']).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=3, value=data['title']).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=4, value=data['name']).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=5, value=data['score']).alignment = Alignment(horizontal='center')


def create_excel_report(results, output_file='bilibili_video_data.xlsx'):
    sorted_results = sorted(results, key=lambda x: x['score'], reverse=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = '视频数据'
    
    headers = ['BVID', '标题', '投稿日期', 'UP主MID', 'UP主名称', '播放量', '弹幕数', '评论数', '收藏数', '投币数', '分享数', '点赞数', '综合评分', '封面路径']
    ws.append(headers)
    apply_header_style(ws, headers)
    
    for i, data in enumerate(sorted_results, start=2):
        row = [
            data['bvid'], data['title'], data['pubdate'], data['mid'], data['name'],
            data['view'], data['danmaku'], data['reply'], data['favorite'], data['coin'],
            data['share'], data['like'], data['score'], data.get('cover_path', '')
        ]
        ws.append(row)
        apply_center_alignment(ws, i, len(headers))
    
    create_rank_section(ws, sorted_results, len(sorted_results) + 1)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
    ws.column_dimensions['B'].width = 40
    
    wb.save(output_file)
    return output_file
