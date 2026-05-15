#!/usr/bin/env python3
import asyncio
import sys
import os
import time
import httpx

sys.path.insert(0, os.path.dirname(__file__))

from backend.db_schema import (
    init_data_dir, get_cache_db, get_chart_db, get_nominations_db,
    create_cache_db, create_chart_db, insert_video_to_chart,
    update_cache_from_chart, get_list_of_charts, create_nominations_db
)
from DataCollector.collector.utils import load_bvids
from DataCollector.collector.data_collector import collect_all_data
from DataCollector.collector.excel_handler import create_excel_report

SERVERCHAN_URL = 'https://7487.push.ft07.com/send/sctp7487t1oar49wlwkey1s7owymazc.send'


def send_serverchan(title, desp):
    try:
        httpx.post(SERVERCHAN_URL, data={'title': title, 'desp': desp}, timeout=10)
    except:
        pass


def print_banner():
    print('''
╔══════════════════════════════════════════╗
║         BiliVBoard 榜单制作工具          ║
╚══════════════════════════════════════════╝
''')


def get_chart_id():
    existing = get_list_of_charts()
    if existing:
        print('已有榜单:')
        for c in existing:
            print(f'  #{c["id"]:03d} | {c["formula"]} | {c["video_count"]} 首')
        print()

    while True:
        try:
            raw = input('请输入本次榜单数字 ID: ').strip()
            chart_id = int(raw)
            if chart_id < 1:
                print('ID 必须 >= 1')
                continue
            db_path = get_chart_db(chart_id)
            if os.path.exists(db_path):
                overwrite = input(f'榜单 #{chart_id:03d} 已存在，覆盖？(y/N): ').strip().lower()
                if overwrite != 'y':
                    continue
            return chart_id
        except ValueError:
            print('请输入数字')


def get_load_from_cache():
    cache_path = get_cache_db()
    if not os.path.exists(cache_path):
        print('缓存数据库不存在，将全量采集')
        return False

    conn = __import__('sqlite3').connect(cache_path)
    count = conn.execute('SELECT COUNT(*) FROM cache').fetchone()[0]
    conn.close()

    if count == 0:
        print('缓存为空，将全量采集')
        return False

    print(f'缓存中有 {count} 首歌曲')
    choice = input('从缓存加载 TOP 1000？(Y/n): ').strip().lower()
    return choice != 'n'


def get_full_mode():
    print('''
采集模式:
  1. 快速模式 (TOP 100)    ~5 分钟
  2. 标准模式 (TOP 500)    ~25 分钟
  3. 完整模式 (全部 10951)  ~12 小时 ⚠️  耗时极长
''')
    while True:
        choice = input('选择模式 (1/2/3，默认1): ').strip()
        if choice == '' or choice == '1':
            return 100
        elif choice == '2':
            return 500
        elif choice == '3':
            confirm = input('确认完整模式？将采集全部 10951 首 (yes/no): ').strip().lower()
            if confirm == 'yes':
                return 99999
            continue
        print('无效选择')


def get_formula():
    print('''
评分公式:
  1. V3 Normal (普通版) - 半衰期约20.79天
  2. V3 Radical (激进版) - 半衰期约6.93天
  3. V3 E SP (无时间衰减) - 适用于长期评估
''')
    while True:
        choice = input('选择公式 (1/2/3，默认1): ').strip()
        if choice == '' or choice == '1':
            return 'normal'
        elif choice == '2':
            return 'radical'
        elif choice == '3':
            return 'e_sp'
        print('无效选择')


async def collect_from_cache(chart_id, formula):
    print('\n从缓存加载 TOP 1000...')
    cache_path = get_cache_db()
    conn = __import__('sqlite3').connect(cache_path)
    conn.row_factory = __import__('sqlite3').Row

    rows = conn.execute('''
        SELECT bvid, title, name, cover_url, best_score AS score,
               best_rank AS rank, view, like_count, coin, favorite,
               danmaku, reply, share
        FROM cache
        WHERE best_score IS NOT NULL
        ORDER BY best_score DESC
        LIMIT 1000
    ''').fetchall()
    conn.close()

    videos = [dict(r) for r in rows]
    print(f'从缓存加载了 {len(videos)} 首')

    db_path = create_chart_db(chart_id, formula)
    conn = __import__('sqlite3').connect(db_path)
    c = conn.cursor()
    for i, v in enumerate(videos, 1):
        bilibili_url = f"https://www.bilibili.com/video/{v['bvid']}"
        c.execute('''
            INSERT OR REPLACE INTO videos (rank, bvid, title, name, view, danmaku,
                reply, favorite, coin, share, like_count, score, cover_url, bilibili_url)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (i, v['bvid'], v['title'], v['name'],
              v.get('view', 0), v.get('danmaku', 0), v.get('reply', 0),
              v.get('favorite', 0), v.get('coin', 0), v.get('share', 0),
              v.get('like_count', 0), v['score'], v.get('cover_url', ''),
              bilibili_url))
    conn.commit()
    conn.close()

    excel_path = os.path.join(os.path.dirname(__file__), 'DataCollector', f'chart_{chart_id:03d}.xlsx')
    _export_excel(videos, excel_path)

    update_cache_from_chart(chart_id, videos)
    return videos


async def collect_full(chart_id, formula, limit):
    print(f'\n开始采集 (限制 {limit} 个)...')

    bvids = load_bvids()
    if limit < len(bvids):
        bvids = bvids[:limit]

    print(f'共 {len(bvids)} 个 BV 号')

    start_time = time.time()
    send_serverchan(f'BiliVBoard #{chart_id:03d} 采集开始', f'采集 {len(bvids)} 个视频\n公式: {formula}')

    results, stats = await collect_all_data(bvids, formula, concurrency=3)

    elapsed = time.time() - start_time
    minutes = int(elapsed // 60)
    seconds = int(elapsed % 60)

    if results:
        sorted_results = sorted(results, key=lambda x: x['score'], reverse=True)

        db_path = create_chart_db(chart_id, formula)
        conn = __import__('sqlite3').connect(db_path)
        c = conn.cursor()
        for i, v in enumerate(sorted_results[:1000], 1):
            bilibili_url = f"https://www.bilibili.com/video/{v['bvid']}"
            c.execute('''
                INSERT OR REPLACE INTO videos (rank, bvid, title, name, view, danmaku,
                    reply, favorite, coin, share, like_count, score, cover_url,
                    bilibili_url, pubdate)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (i, v['bvid'], v['title'], v['name'],
                  v.get('view', 0), v.get('danmaku', 0), v.get('reply', 0),
                  v.get('favorite', 0), v.get('coin', 0), v.get('share', 0),
                  v.get('like_count', 0), v['score'], v.get('cover_url', ''),
                  bilibili_url, v.get('pubdate', '')))
        conn.commit()
        conn.close()

        excel_path = os.path.join(os.path.dirname(__file__), 'DataCollector', f'chart_{chart_id:03d}.xlsx')
        _export_excel(sorted_results[:1000], excel_path)

        update_cache_from_chart(chart_id, sorted_results[:1000])

    title = f'BiliVBoard #{chart_id:03d} 采集完成'
    desp = f"""## 采集结果

- **榜单**: #{chart_id:03d}
- **总数**: {len(bvids)}
- **成功**: {len(results)}
- **失败**: {stats['failed']}
- **耗时**: {minutes}分{seconds}秒
- **公式**: {formula}
"""
    send_serverchan(title, desp)
    print(f'\n完成! 耗时: {minutes}分{seconds}秒')
    return results


def _export_excel(videos, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = '排行榜'

    headers = ['排名', 'BVID', '标题', 'UP主', '评分', '播放', '点赞', '投币', '收藏', '弹幕', '评论', '转发']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for i, v in enumerate(videos, 2):
        ws.cell(row=i, column=1, value=i-1)
        ws.cell(row=i, column=2, value=v.get('bvid', ''))
        ws.cell(row=i, column=3, value=v.get('title', ''))
        ws.cell(row=i, column=4, value=v.get('name', ''))
        ws.cell(row=i, column=5, value=v.get('score', 0))
        ws.cell(row=i, column=6, value=v.get('view', 0))
        ws.cell(row=i, column=7, value=v.get('like_count', 0))
        ws.cell(row=i, column=8, value=v.get('coin', 0))
        ws.cell(row=i, column=9, value=v.get('favorite', 0))
        ws.cell(row=i, column=10, value=v.get('danmaku', 0))
        ws.cell(row=i, column=11, value=v.get('reply', 0))
        ws.cell(row=i, column=12, value=v.get('share', 0))
        for col in range(1, 13):
            ws.cell(row=i, column=col).alignment = Alignment(horizontal='center')

    for col in range(1, 13):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
    ws.column_dimensions['C'].width = 40

    wb.save(path)
    print(f'Excel: {path}')


async def main():
    print_banner()
    init_data_dir()

    create_cache_db()
    create_nominations_db()

    chart_id = get_chart_id()
    load_cache = get_load_from_cache()
    formula = get_formula()

    if load_cache:
        videos = await collect_from_cache(chart_id, formula)
    else:
        limit = get_full_mode()
        videos = await collect_full(chart_id, formula, limit)

    print(f'\n榜单 #{chart_id:03d} 制作完成!')
    print(f'数据库: {get_chart_db(chart_id)}')


if __name__ == '__main__':
    asyncio.run(main())
